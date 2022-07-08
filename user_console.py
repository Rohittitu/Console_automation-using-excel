# -*- coding: utf-8 -*-
"""
Created on Wed Jul  6 20:54:44 2022

@author: RSuman
"""


import socket
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import Progressbar
import serial, time
import serial.tools.list_ports
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
import threading

def pop_socket():
    global pp
    pp = socket.socket()         # Create a socket object
    host = socket.gethostname() # Get local machine name
    port = 12345              # Reserve a port for your service.
    pp.connect((host, port))
    print (pp.recv(1024))
    pp.send("Internal socket has been created \n".encode())
pop_socket()
time.sleep(5)
def send_for_popup(popup_information):
    pp.send(popup_information.encode())
    return (pp.recv(1024).decode())


def info(count):
    if (sheet.cell(row=count, column=6).value != None):
        info_val = send_for_popup("info" + sheet.cell(row=count, column=6).value)
        if (info_val == "ok"):
            return True
        else:
            return False
    else:
        return True

def tester_response(count):
    if (sheet.cell(row=count, column=7).value != None):
        quest_Ans = send_for_popup("yes_no" + sheet.cell(row=count, column=7).value)
        sheet.cell(row=count, column=8).value = quest_Ans
        if quest_Ans == "Yes":
            return True
        else:
            return False
    else:
        return True



def load_excel(excel_name):
    global sheet
    global wb
    wb = openpyxl.load_workbook(excel_name)
    sheet = wb["Test_case"]
    print("Test details has been imported")


ports = list(serial.tools.list_ports.comports())
available_com= []
for p in ports:
    _com = str(p).split(" ")[0]
    available_com.append(_com)

class Can_selection:

    def __init__(self, master):
        self.var_ent = StringVar()
        self.var_ent.set("Please select the file")
        self.sigsinit_flag = False
        self.comm= StringVar()
        self.comm.set("COM")
        #self.all_com = StringVar()
        self.all_com = available_com


        self.bound_ = [110, 300, 600, 1200, 2400, 4800, 9600, 14400, 19200, 38400, 57600, 115200,230400, 460800, 921600]
        self.bound_rate = IntVar()
        self.bound_rate.set(000)

        self.can_se = Frame(master, height=250, width=550)
        self.can_se.pack()

        self.labl1  = Label(self.can_se, text= "Port:",font=('Helvetica', 10,'bold'))
        self.com_selction = OptionMenu(self.can_se,self.comm, *self.all_com)
        self.labl2  = Label(self.can_se, text= "Bound Rate:",font=('Helvetica', 10,'bold'))
        self.bound_selction = OptionMenu(self.can_se, self.bound_rate, * self.bound_ )
        self.brow   = Entry(self.can_se, bd =3,textvariable= self.var_ent, font=('Helvetica', 10),width=50)
        self.brow_b = Button(self.can_se, text ="BROWSE", bg='gray69',width=10,bd=4,relief='ridge', font=('Helvetica', 10,'bold'),height=1,command =self.brws_sig)
        self.connect = Button(self.can_se, text="Connect", bg='gray69', width=10, bd=4, relief='ridge',
                             font=('Helvetica', 10, 'bold'), height=1, command=self.establish_connection)
        self.brow_s = Button(self.can_se, text="Start", bg='gray69', width=10, bd=4, relief='ridge',
                             font=('Helvetica', 10, 'bold'), height=1, command= self.start_execution)
        self.Overall_prog = Progressbar(self.can_se, orient=HORIZONTAL, length=530, mode='determinate')

        self.labl1.place(x=30, y=10)
        self.com_selction.place(x=90, y=10)
        self.connect.place(x=400, y=10)
        self.labl2.place(x=200, y=10)
        self.bound_selction.place(x=300, y=10)
        self.brow.place(x= 10, y= 60)
        self.brow_b.place(x= 400, y= 55)
        self.brow_s.place(x=200, y=100)
        self.Overall_prog.place(x=10, y=200)


    def brws_sig(self):
        toplevel = Tk()
        toplevel.withdraw()
        global sig_name
        sig_name = filedialog.askopenfilename()
        if sig_name:
            self.var_ent.set(sig_name)
            print(sig_name)
            self.sigsinit_flag = True
            load_excel(sig_name)
        else:
            print("No file has been selected")
            self.sigsinit_flag = False


    def establish_connection(self):
        try:
            self.ser = serial.Serial(self.comm.get(), self.bound_rate.get())
        except Exception as e:
            print(e)
        else:
            print("connection has been establieshed for port: " + self.comm.get() + "with bound rate of: " + str(self.bound_rate.get()))

    def send_command(self, cmmd):
        temp_cmd = cmmd + '\n'
        if self.ser.isOpen():
            self.ser.write(temp_cmd.encode())

    def get_responses(self):
        buffer = b''
        while True:
            buffer = buffer + self.ser.read(self.ser.inWaiting())
            time.sleep(0.5)
            if self.ser.inWaiting() == 0:
                break
        return buffer

    def execute(self,cmd):
        self.send_command(cmd)
        response = self.get_responses()
        return response.decode("utf-8")


    def start(self):
        if self.sigsinit_flag == True:

            for testcase in range(2, sheet.max_row + 1):
                info(testcase)
                if (sheet.cell(row=testcase, column=2).value != None):
                    resp= self.execute(sheet.cell(row=testcase, column=2).value)
                    try:
                        sheet.cell(row=testcase, column=3).value = ILLEGAL_CHARACTERS_RE.sub(r'',resp)

                    except Exception as ee:
                        print(ee)
                        sheet.cell(row=testcase, column=3).value= ee

                    prog = ((testcase + 1) * 100) // (sheet.max_row + 1)

                    self.Overall_prog["value"] = prog
                    self.Overall_prog.update()

                    if sheet.cell(row=testcase, column=5).value == None:
                        wb.save("report.xlsx")
                        time.sleep(1)
                    else:

                        if sheet.cell(row=testcase, column=5).value in resp:
                            if tester_response(testcase) == True:
                                sheet.cell(row=testcase, column=4).value = 'PASSED'
                                sheet.cell(row=testcase, column=4).fill = PatternFill(start_color='5CEF3B', end_color='5CEF3B',fill_type='solid')
                                wb.save("report.xlsx")

                            else:
                                sheet.cell(row=testcase, column=4).value = 'FAILED'
                                sheet.cell(row=testcase, column=4).fill = PatternFill(start_color='F2421A',end_color='F2421A',fill_type='solid')
                                wb.save("report.xlsx")
                        else:
                            sheet.cell(row=testcase, column=4).value= 'FAILED'
                            sheet.cell(row=testcase, column=4).fill = PatternFill(start_color='F2421A', end_color='F2421A', fill_type = 'solid')
                            wb.save("report.xlsx")

            print("Execution Completed")


        else:
            print("No file has been imported")

    def start_execution(self):
        threading.Thread(target=self.start).start()


root = Tk()
root.title("Console Automation ")
app = Can_selection(root)
root.mainloop()